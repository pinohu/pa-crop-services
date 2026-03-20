from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ========== SHEET 1: ASSUMPTIONS ==========
ws1 = wb.active
ws1.title = "Assumptions"

blue_font = Font(name='Arial', size=11, color='0000FF')
black_font = Font(name='Arial', size=11)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
section_font = Font(name='Arial', size=12, bold=True)
title_font = Font(name='Arial', size=14, bold=True)
pct_fmt = '0.0%'
curr_fmt = '$#,##0'
num_fmt = '#,##0'
header_fill = PatternFill('solid', fgColor='534AB7')
input_fill = PatternFill('solid', fgColor='FFFF00')
section_fill = PatternFill('solid', fgColor='E8E6F8')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_range(ws, row, col, val, font=black_font, fill=None, fmt=None, width=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='left' if isinstance(val, str) else 'right')
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    return cell

# Title
ws1.cell(row=1, column=1, value="PA CROP Business — Financial Assumptions").font = title_font
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 30

# STARTUP COSTS
r = 3
style_range(ws1, r, 1, "STARTUP COSTS", section_font, section_fill)
style_range(ws1, r, 2, "Amount", section_font, section_fill)
style_range(ws1, r, 3, "Notes", section_font, section_fill)
items = [
    ("PA LLC filing", 125, "One-time, PA DOS"),
    ("CROP registration", 150, "Statement of address filing"),
    ("Domain registration", 15, "Annual, pacropservices.com"),
    ("Professional liability insurance", 750, "Annual E&O policy"),
    ("Legal review of service agreement", 750, "Attorney review"),
    ("Initial Google Ads budget", 500, "Month 1 testing"),
    ("Office supplies (scanner, etc.)", 350, "One-time"),
    ("Marketing materials / brochures", 200, "Partner outreach collateral"),
    ("Technology stack", 0, "Existing Dynasty licenses — $0 marginal"),
]
for i, (label, val, note) in enumerate(items):
    style_range(ws1, r+1+i, 1, label)
    style_range(ws1, r+1+i, 2, val, blue_font, input_fill, curr_fmt)
    style_range(ws1, r+1+i, 3, note)
total_row = r + 1 + len(items)
style_range(ws1, total_row, 1, "TOTAL STARTUP", Font(name='Arial', size=11, bold=True))
style_range(ws1, total_row, 2, None, Font(name='Arial', size=11, bold=True))
ws1.cell(row=total_row, column=2).value = f'=SUM(B{r+1}:B{total_row-1})'
ws1.cell(row=total_row, column=2).number_format = curr_fmt

# PRICING ASSUMPTIONS
r2 = total_row + 2
style_range(ws1, r2, 1, "PRICING TIERS", section_font, section_fill)
style_range(ws1, r2, 2, "Annual Price", section_font, section_fill)
style_range(ws1, r2, 3, "Description", section_font, section_fill)
tiers = [
    ("Starter", 79, "Basic CROP + mail forwarding"),
    ("Professional", 179, "CROP + scanning + compliance calendar"),
    ("Premium", 299, "Full service + filing + notary"),
    ("Partner (white-label)", 99, "Per-client, CPA/attorney channel"),
    ("Annual report filing add-on", 75, "Per filing, on top of $7 state fee"),
    ("LLC formation package", 650, "Includes $125 state fee"),
    ("Notarization", 25, "Per document"),
    ("Virtual mailbox add-on", 180, "Per year ($15/month)"),
    ("Compliance monitoring add-on", 120, "Per year ($10/month)"),
]
for i, (label, price, desc) in enumerate(tiers):
    style_range(ws1, r2+1+i, 1, label)
    style_range(ws1, r2+1+i, 2, price, blue_font, input_fill, curr_fmt)
    style_range(ws1, r2+1+i, 3, desc)

# OPERATING COST ASSUMPTIONS
r3 = r2 + 1 + len(tiers) + 1
style_range(ws1, r3, 1, "MONTHLY OPERATING COSTS", section_font, section_fill)
style_range(ws1, r3, 2, "Monthly", section_font, section_fill)
style_range(ws1, r3, 3, "Notes", section_font, section_fill)
ops = [
    ("Office overhead (allocated)", 200, "Portion of existing Erie office"),
    ("Google Ads", 500, "Scaling with revenue"),
    ("Postage and mail forwarding", 100, "Scales with client count"),
    ("Part-time staff (mail handling)", 500, "Needed at 200+ clients"),
    ("Insurance (monthly)", 65, "E&O amortized"),
    ("Miscellaneous", 100, "Supplies, printing"),
    ("Stripe processing (est.)", 300, "~3% of monthly revenue"),
]
for i, (label, val, note) in enumerate(ops):
    style_range(ws1, r3+1+i, 1, label)
    style_range(ws1, r3+1+i, 2, val, blue_font, input_fill, curr_fmt)
    style_range(ws1, r3+1+i, 3, note)
ops_total = r3 + 1 + len(ops)
style_range(ws1, ops_total, 1, "TOTAL MONTHLY OPEX", Font(name='Arial', size=11, bold=True))
ws1.cell(row=ops_total, column=2, value=f'=SUM(B{r3+1}:B{ops_total-1})').number_format = curr_fmt
ws1.cell(row=ops_total, column=2).font = Font(name='Arial', size=11, bold=True)

# GROWTH ASSUMPTIONS
r4 = ops_total + 2
style_range(ws1, r4, 1, "GROWTH ASSUMPTIONS", section_font, section_fill)
style_range(ws1, r4, 2, "Value", section_font, section_fill)
growth = [
    ("New clients per month (Month 1-3)", 15, num_fmt),
    ("New clients per month (Month 4-6)", 40, num_fmt),
    ("New clients per month (Month 7-12)", 75, num_fmt),
    ("New clients per month (Month 13-18)", 125, num_fmt),
    ("New clients per month (Month 19-24)", 175, num_fmt),
    ("Monthly churn rate", 0.005, pct_fmt),
    ("Starter tier mix", 0.30, pct_fmt),
    ("Professional tier mix", 0.40, pct_fmt),
    ("Premium tier mix", 0.10, pct_fmt),
    ("Partner tier mix", 0.20, pct_fmt),
    ("Add-on adoption rate", 0.25, pct_fmt),
    ("Average add-on revenue per adopter", 75, curr_fmt),
]
for i, (label, val, fmt) in enumerate(growth):
    style_range(ws1, r4+1+i, 1, label)
    style_range(ws1, r4+1+i, 2, val, blue_font, input_fill, fmt)

# ========== SHEET 2: 24-MONTH P&L ==========
ws2 = wb.create_sheet("24-Month P&L")
ws2.cell(row=1, column=1, value="PA CROP Business — 24-Month Profit & Loss").font = title_font
ws2.column_dimensions['A'].width = 32

months = list(range(1, 25))
for i, m in enumerate(months):
    col = i + 2
    ws2.column_dimensions[get_column_letter(col)].width = 12
    c = ws2.cell(row=3, column=col, value=f"Month {m}")
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

ws2.cell(row=3, column=1, value="").font = header_font

# Row labels
labels = [
    ("REVENUE", True),
    ("New clients acquired", False),
    ("Churned clients", False),
    ("Total active clients", False),
    ("Starter revenue", False),
    ("Professional revenue", False),
    ("Premium revenue", False),
    ("Partner revenue", False),
    ("Add-on revenue", False),
    ("TOTAL REVENUE", True),
    ("", False),
    ("EXPENSES", True),
    ("Office overhead", False),
    ("Google Ads", False),
    ("Postage/mail", False),
    ("Staff", False),
    ("Insurance", False),
    ("Miscellaneous", False),
    ("Stripe fees (3%)", False),
    ("TOTAL EXPENSES", True),
    ("", False),
    ("NET PROFIT", True),
    ("Cumulative profit", False),
    ("", False),
    ("KEY METRICS", True),
    ("MRR", False),
    ("ARR run rate", False),
    ("Avg revenue per client", False),
    ("Gross margin", False),
]

for i, (label, is_bold) in enumerate(labels):
    r = 4 + i
    c = ws2.cell(row=r, column=1, value=label)
    c.font = Font(name='Arial', size=11, bold=is_bold)
    c.border = thin_border
    if is_bold and label:
        c.fill = PatternFill('solid', fgColor='E8E6F8')

# Populate formulas for each month
for mi, m in enumerate(months):
    col = mi + 2
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1) if mi > 0 else None

    # New clients (hardcoded based on growth assumptions)
    if m <= 3: new_clients = 15
    elif m <= 6: new_clients = 40
    elif m <= 12: new_clients = 75
    elif m <= 18: new_clients = 125
    else: new_clients = 175

    # Row 4: blank (REVENUE header)
    # Row 5: New clients
    ws2.cell(row=5, column=col, value=new_clients).font = blue_font
    ws2.cell(row=5, column=col).number_format = num_fmt
    ws2.cell(row=5, column=col).border = thin_border

    # Row 6: Churned (0.5% of prior total, 0 for month 1)
    if mi == 0:
        ws2.cell(row=6, column=col, value=0).border = thin_border
    else:
        ws2.cell(row=6, column=col, value=f'=ROUND({prev}7*0.005,0)').border = thin_border
    ws2.cell(row=6, column=col).number_format = num_fmt

    # Row 7: Total active = prior total + new - churned (or just new for month 1)
    if mi == 0:
        ws2.cell(row=7, column=col, value=f'={cl}5-{cl}6').border = thin_border
    else:
        ws2.cell(row=7, column=col, value=f'={prev}7+{cl}5-{cl}6').border = thin_border
    ws2.cell(row=7, column=col).number_format = num_fmt
    ws2.cell(row=7, column=col).font = Font(name='Arial', size=11, bold=True)

    # Row 8: Starter revenue (30% of clients * $79/12)
    ws2.cell(row=8, column=col, value=f'=ROUND({cl}7*0.30*79/12,0)').border = thin_border
    ws2.cell(row=8, column=col).number_format = curr_fmt

    # Row 9: Professional revenue (40% * $179/12)
    ws2.cell(row=9, column=col, value=f'=ROUND({cl}7*0.40*179/12,0)').border = thin_border
    ws2.cell(row=9, column=col).number_format = curr_fmt

    # Row 10: Premium revenue (10% * $299/12)
    ws2.cell(row=10, column=col, value=f'=ROUND({cl}7*0.10*299/12,0)').border = thin_border
    ws2.cell(row=10, column=col).number_format = curr_fmt

    # Row 11: Partner revenue (20% * $99/12)
    ws2.cell(row=11, column=col, value=f'=ROUND({cl}7*0.20*99/12,0)').border = thin_border
    ws2.cell(row=11, column=col).number_format = curr_fmt

    # Row 12: Add-on revenue (25% of clients * $75/12)
    ws2.cell(row=12, column=col, value=f'=ROUND({cl}7*0.25*75/12,0)').border = thin_border
    ws2.cell(row=12, column=col).number_format = curr_fmt

    # Row 13: Total Revenue
    ws2.cell(row=13, column=col, value=f'=SUM({cl}8:{cl}12)').border = thin_border
    ws2.cell(row=13, column=col).number_format = curr_fmt
    ws2.cell(row=13, column=col).font = Font(name='Arial', size=11, bold=True)

    # Row 14: blank
    # EXPENSES (rows 15-21)
    # Row 16: Office
    ws2.cell(row=16, column=col, value=200).border = thin_border
    ws2.cell(row=16, column=col).number_format = curr_fmt
    # Row 17: Ads (scales: $300 m1-3, $500 m4-12, $750 m13+)
    if m <= 3: ads = 300
    elif m <= 12: ads = 500
    else: ads = 750
    ws2.cell(row=17, column=col, value=ads).border = thin_border
    ws2.cell(row=17, column=col).number_format = curr_fmt
    # Row 18: Postage (scales with clients)
    ws2.cell(row=18, column=col, value=f'=ROUND({cl}7*0.15,0)').border = thin_border
    ws2.cell(row=18, column=col).number_format = curr_fmt
    # Row 19: Staff (0 until 200 clients, then scales)
    ws2.cell(row=19, column=col, value=f'=IF({cl}7<200,0,IF({cl}7<500,500,IF({cl}7<1000,2000,4000)))').border = thin_border
    ws2.cell(row=19, column=col).number_format = curr_fmt
    # Row 20: Insurance
    ws2.cell(row=20, column=col, value=65).border = thin_border
    ws2.cell(row=20, column=col).number_format = curr_fmt
    # Row 21: Misc
    ws2.cell(row=21, column=col, value=100).border = thin_border
    ws2.cell(row=21, column=col).number_format = curr_fmt
    # Row 22: Stripe (3% of revenue)
    ws2.cell(row=22, column=col, value=f'=ROUND({cl}13*0.03,0)').border = thin_border
    ws2.cell(row=22, column=col).number_format = curr_fmt
    # Row 23: Total expenses
    ws2.cell(row=23, column=col, value=f'=SUM({cl}16:{cl}22)').border = thin_border
    ws2.cell(row=23, column=col).number_format = curr_fmt
    ws2.cell(row=23, column=col).font = Font(name='Arial', size=11, bold=True)

    # Row 25: Net Profit
    ws2.cell(row=25, column=col, value=f'={cl}13-{cl}23').border = thin_border
    ws2.cell(row=25, column=col).number_format = '$#,##0;($#,##0);-'
    ws2.cell(row=25, column=col).font = Font(name='Arial', size=11, bold=True)

    # Row 26: Cumulative
    if mi == 0:
        ws2.cell(row=26, column=col, value=f'={cl}25-Assumptions!B{total_row}').border = thin_border
    else:
        ws2.cell(row=26, column=col, value=f'={prev}26+{cl}25').border = thin_border
    ws2.cell(row=26, column=col).number_format = '$#,##0;($#,##0);-'

    # KEY METRICS (rows 28-31)
    # Row 29: MRR
    ws2.cell(row=29, column=col, value=f'={cl}13').border = thin_border
    ws2.cell(row=29, column=col).number_format = curr_fmt
    # Row 30: ARR
    ws2.cell(row=30, column=col, value=f'={cl}13*12').border = thin_border
    ws2.cell(row=30, column=col).number_format = curr_fmt
    ws2.cell(row=30, column=col).font = Font(name='Arial', size=11, bold=True)
    # Row 31: Avg rev/client
    ws2.cell(row=31, column=col, value=f'=IF({cl}7>0,{cl}13/{cl}7*12,0)').border = thin_border
    ws2.cell(row=31, column=col).number_format = '$#,##0'
    # Row 32: Gross margin
    ws2.cell(row=32, column=col, value=f'=IF({cl}13>0,({cl}13-{cl}23)/{cl}13,0)').border = thin_border
    ws2.cell(row=32, column=col).number_format = pct_fmt

# ========== SHEET 3: UNIT ECONOMICS ==========
ws3 = wb.create_sheet("Unit Economics")
ws3.cell(row=1, column=1, value="PA CROP Business — Unit Economics").font = title_font
ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 40

metrics = [
    ("", "", ""),
    ("ACQUISITION METRICS", "", ""),
    ("Customer Acquisition Cost (blended)", 25, "Weighted avg across all channels"),
    ("CAC — SEO / Organic", 10, "Content cost amortized over leads"),
    ("CAC — Google Ads", 45, "$500/mo budget, ~11 clients/mo"),
    ("CAC — Partner channel", 15, "One-time setup per partner firm"),
    ("CAC — Direct outreach", 8, "Email/LinkedIn, mostly time cost"),
    ("CAC — Referral", 20, "$20 referral bonus"),
    ("", "", ""),
    ("LIFETIME VALUE", "", ""),
    ("Average annual revenue per client", 150, "Blended across all tiers + add-ons"),
    ("Average client lifespan (years)", 5, "CROP services are sticky — low churn"),
    ("Customer Lifetime Value (LTV)", 750, "=5 years x $150/yr"),
    ("LTV:CAC Ratio", "30:1", "Target: >3:1"),
    ("Payback period (months)", 2, "Revenue covers CAC in month 2"),
    ("", "", ""),
    ("PROFITABILITY METRICS", "", ""),
    ("Gross margin per client", "87%", "Revenue minus direct costs"),
    ("Operating cost per client/year", 20, "Total opex / total clients"),
    ("Net profit per client/year", 130, "$150 revenue - $20 opex"),
    ("Human labor per client/year", "15 min", "Fully automated except mail handling"),
    ("", "", ""),
    ("SCALE ECONOMICS", "", ""),
    ("Clients needed for $100K ARR", 667, "$100K / $150 avg"),
    ("Clients needed for $300K ARR", 2000, "$300K / $150 avg"),
    ("Clients needed for $600K ARR", 4000, "$600K / $150 avg"),
    ("Partner firms needed for $200K ARR", 27, "@ 75 clients/firm x $99"),
    ("", "", ""),
    ("EXIT VALUATION", "", ""),
    ("Revenue multiple (low)", "3.0x", "Compliance businesses trade at 3-5x ARR"),
    ("Revenue multiple (high)", "5.0x", "High-margin, recurring, low-churn premium"),
    ("Valuation at $300K ARR (low)", "$900K", ""),
    ("Valuation at $300K ARR (high)", "$1.5M", ""),
    ("Valuation at $600K ARR (low)", "$1.8M", ""),
    ("Valuation at $600K ARR (high)", "$3.0M", ""),
]

for i, (label, val, note) in enumerate(metrics):
    r = 2 + i
    c1 = ws3.cell(row=r, column=1, value=label)
    c2 = ws3.cell(row=r, column=2, value=val)
    c3 = ws3.cell(row=r, column=3, value=note)
    if label and val == "" and note == "":
        c1.font = section_font
        c1.fill = section_fill
        c2.fill = section_fill
        c3.fill = section_fill
    else:
        c1.font = Font(name='Arial', size=11)
        c2.font = blue_font if isinstance(val, (int, float)) else Font(name='Arial', size=11)
        c3.font = Font(name='Arial', size=10, color='666666')
    for c in [c1, c2, c3]:
        c.border = thin_border

# ========== SHEET 4: SCENARIO ANALYSIS ==========
ws4 = wb.create_sheet("Scenarios")
ws4.cell(row=1, column=1, value="PA CROP Business — Scenario Analysis").font = title_font
ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18

scenarios = [
    ("", "Conservative", "Base Case", "Aggressive"),
    ("YEAR 1 PROJECTIONS", "", "", ""),
    ("Total clients (end of Year 1)", 400, 800, 1500),
    ("Blended avg revenue/client", 130, 150, 170),
    ("Year 1 ARR", "=B4*B5", "=C4*C5", "=D4*D5"),
    ("Total Year 1 revenue", "=B6*0.6", "=C6*0.65", "=D6*0.7"),
    ("Total Year 1 expenses", 18000, 22000, 35000),
    ("Year 1 net profit", "=B7-B8", "=C7-C8", "=D7-D8"),
    ("", "", "", ""),
    ("YEAR 2 PROJECTIONS", "", "", ""),
    ("Total clients (end of Year 2)", 1200, 2500, 5000),
    ("Blended avg revenue/client", 140, 155, 175),
    ("Year 2 ARR", "=B12*B13", "=C12*C13", "=D12*D13"),
    ("Total Year 2 revenue", "=B14*0.85", "=C14*0.9", "=D14*0.95"),
    ("Total Year 2 expenses", 36000, 55000, 95000),
    ("Year 2 net profit", "=B15-B16", "=C15-C16", "=D15-D16"),
    ("", "", "", ""),
    ("CUMULATIVE (2 YEARS)", "", "", ""),
    ("Total revenue", "=B7+B15", "=C7+C15", "=D7+D15"),
    ("Total expenses", "=B8+B16", "=C8+C16", "=D8+D16"),
    ("Total net profit", "=B19-B20", "=C19-C20", "=D19-D20"),
    ("ROI on startup investment", "=B21/2840", "=C21/2840", "=D21/2840"),
    ("", "", "", ""),
    ("EXIT VALUATION (at Year 2 ARR)", "", "", ""),
    ("At 3x ARR", "=B14*3", "=C14*3", "=D14*3"),
    ("At 5x ARR", "=B14*5", "=C14*5", "=D14*5"),
]

for i, row_data in enumerate(scenarios):
    r = 3 + i
    for j, val in enumerate(row_data):
        cell = ws4.cell(row=r, column=j+1, value=val)
        cell.border = thin_border
        if i == 0:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        elif isinstance(val, str) and val and not val.startswith('=') and j == 0:
            if val.isupper() or val.startswith("YEAR") or val.startswith("CUM") or val.startswith("EXIT"):
                cell.font = section_font
                cell.fill = section_fill
            else:
                cell.font = Font(name='Arial', size=11)
        elif isinstance(val, (int, float)):
            cell.font = blue_font
            cell.fill = input_fill
            if any(x in scenarios[i][0].lower() for x in ['revenue', 'expense', 'profit', 'arr', 'valuation', 'arR']):
                cell.number_format = curr_fmt
            elif 'roi' in scenarios[i][0].lower():
                cell.number_format = '0%'
            else:
                cell.number_format = num_fmt
        elif isinstance(val, str) and val.startswith('='):
            cell.font = Font(name='Arial', size=11)
            if any(x in scenarios[i][0].lower() for x in ['revenue', 'expense', 'profit', 'arr', 'valuation']):
                cell.number_format = curr_fmt
            elif 'roi' in scenarios[i][0].lower():
                cell.number_format = '0%'

wb.save('/home/claude/PA-CROP-Financial-Model.xlsx')
print("Financial model created")
